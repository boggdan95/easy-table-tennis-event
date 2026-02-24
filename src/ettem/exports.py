"""
Export Module for ETTEM
Generates Excel workbooks and CSV exports for tournament data.
"""

import io
import csv
import json
from typing import List, Dict, Any, Optional
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _style_header_row(ws, num_cols: int):
    """Apply consistent header styling to the first row."""
    if not HAS_OPENPYXL:
        return
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def _auto_width(ws):
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 3, 40)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 8)


def generate_tournament_excel(
    tournament_name: str,
    players: List[Dict[str, Any]],
    groups: List[Dict[str, Any]],
    group_matches: List[Dict[str, Any]],
    standings: List[Dict[str, Any]],
    bracket_matches: List[Dict[str, Any]],
    final_positions: List[Dict[str, Any]],
    branding: Optional[Dict[str, Any]] = None,
) -> bytes:
    """
    Generate a multi-sheet Excel workbook with all tournament data.

    Returns: Excel file as bytes.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. pip install openpyxl")

    wb = Workbook()

    # --- Sheet 1: Players ---
    ws_players = wb.active
    ws_players.title = "Jugadores"
    headers = ["#", "Nombre", "Apellido", "País", "Categoría", "Ranking", "Grupo", "Seed"]
    ws_players.append(headers)
    _style_header_row(ws_players, len(headers))
    for i, p in enumerate(players, 1):
        ws_players.append([
            i,
            p.get("nombre", ""),
            p.get("apellido", ""),
            p.get("pais_cd", ""),
            p.get("categoria", ""),
            p.get("ranking_pts", 0),
            p.get("group_name", ""),
            p.get("group_number", ""),
        ])
    _auto_width(ws_players)

    # --- Sheet 2: Groups ---
    ws_groups = wb.create_sheet("Grupos")
    headers = ["Categoría", "Grupo", "Jugadores"]
    ws_groups.append(headers)
    _style_header_row(ws_groups, len(headers))
    for g in groups:
        ws_groups.append([
            g.get("category", ""),
            g.get("name", ""),
            g.get("num_players", 0),
        ])
    _auto_width(ws_groups)

    # --- Sheet 3: Group Results ---
    ws_results = wb.create_sheet("Resultados Grupo")
    headers = ["Categoría", "Grupo", "#", "Jugador 1", "Jugador 2", "Ganador", "Sets", "Estado"]
    ws_results.append(headers)
    _style_header_row(ws_results, len(headers))
    for m in group_matches:
        ws_results.append([
            m.get("category", ""),
            m.get("group_name", ""),
            m.get("match_order", ""),
            m.get("player1_name", ""),
            m.get("player2_name", ""),
            m.get("winner_name", "-"),
            m.get("sets_result", "-"),
            m.get("status", ""),
        ])
    _auto_width(ws_results)

    # --- Sheet 4: Standings ---
    ws_standings = wb.create_sheet("Clasificación")
    headers = ["Categoría", "Grupo", "Pos", "Jugador", "País", "Puntos", "V", "D", "Sets+", "Sets-", "Pts+", "Pts-"]
    ws_standings.append(headers)
    _style_header_row(ws_standings, len(headers))

    # Highlight rows for qualified positions (1st, 2nd)
    gold_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    for s in standings:
        row = [
            s.get("category", ""),
            s.get("group_name", ""),
            s.get("position", ""),
            s.get("player_name", ""),
            s.get("pais_cd", ""),
            s.get("points_total", 0),
            s.get("wins", 0),
            s.get("losses", 0),
            s.get("sets_w", 0),
            s.get("sets_l", 0),
            s.get("points_w", 0),
            s.get("points_l", 0),
        ]
        ws_standings.append(row)
        if s.get("position", 99) <= 2:
            for col in range(1, len(headers) + 1):
                ws_standings.cell(row=ws_standings.max_row, column=col).fill = gold_fill
    _auto_width(ws_standings)

    # --- Sheet 5: Bracket ---
    ws_bracket = wb.create_sheet("Bracket")
    headers = ["Categoría", "Ronda", "#", "Jugador 1", "Jugador 2", "Ganador", "Sets", "Estado"]
    ws_bracket.append(headers)
    _style_header_row(ws_bracket, len(headers))
    for m in bracket_matches:
        ws_bracket.append([
            m.get("category", ""),
            m.get("round_name", ""),
            m.get("match_order", ""),
            m.get("player1_name", "TBD"),
            m.get("player2_name", "TBD"),
            m.get("winner_name", "-"),
            m.get("sets_result", "-"),
            m.get("status", ""),
        ])
    _auto_width(ws_bracket)

    # --- Sheet 6: Final Positions ---
    ws_final = wb.create_sheet("Posiciones Finales")
    headers = ["Categoría", "Posición", "Jugador", "País"]
    ws_final.append(headers)
    _style_header_row(ws_final, len(headers))

    medal_fills = {
        1: PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
        2: PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"),
        3: PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid"),
    }
    for fp in final_positions:
        pos = fp.get("position", 99)
        ws_final.append([
            fp.get("category", ""),
            pos,
            fp.get("player_name", ""),
            fp.get("pais_cd", ""),
        ])
        if pos in medal_fills:
            for col in range(1, len(headers) + 1):
                ws_final.cell(row=ws_final.max_row, column=col).fill = medal_fills[pos]
    _auto_width(ws_final)

    # --- Add tournament info at top of first sheet ---
    if branding:
        # Insert info rows at top of Players sheet
        ws_players.insert_rows(1, 3)
        ws_players.cell(row=1, column=1).value = branding.get("official_name", tournament_name) or tournament_name
        ws_players.cell(row=1, column=1).font = Font(bold=True, size=14)
        if branding.get("organizer"):
            ws_players.cell(row=2, column=1).value = branding["organizer"]
            ws_players.cell(row=2, column=1).font = Font(size=11, color="666666")
        ws_players.cell(row=3, column=1).value = f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_players.cell(row=3, column=1).font = Font(size=9, color="999999")

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def generate_results_csv(
    group_matches: List[Dict[str, Any]],
    bracket_matches: List[Dict[str, Any]],
) -> bytes:
    """Generate a combined CSV of all results (groups + bracket)."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["Fase", "Categoría", "Grupo/Ronda", "#", "Jugador 1", "Jugador 2", "Ganador", "Sets", "Estado"])

    for m in group_matches:
        writer.writerow([
            "Grupos",
            m.get("category", ""),
            m.get("group_name", ""),
            m.get("match_order", ""),
            m.get("player1_name", ""),
            m.get("player2_name", ""),
            m.get("winner_name", "-"),
            m.get("sets_result", "-"),
            m.get("status", ""),
        ])

    for m in bracket_matches:
        writer.writerow([
            "Bracket",
            m.get("category", ""),
            m.get("round_name", ""),
            m.get("match_order", ""),
            m.get("player1_name", "TBD"),
            m.get("player2_name", "TBD"),
            m.get("winner_name", "-"),
            m.get("sets_result", "-"),
            m.get("status", ""),
        ])

    csv_content = output.getvalue()
    return ('\ufeff' + csv_content).encode('utf-8')
